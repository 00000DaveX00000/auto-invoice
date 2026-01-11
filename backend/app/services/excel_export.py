from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from ..schemas.invoice import InvoiceResponse, VoucherEntry, CategorySummary


def create_invoice_excel(
    invoices: List[InvoiceResponse],
    summary: List[CategorySummary],
    anomalies: List[InvoiceResponse],
    vouchers: List[VoucherEntry],
) -> BytesIO:
    """创建包含4个Sheet的Excel文件"""
    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sheet 1: 发票明细表
    ws1 = wb.active
    ws1.title = "发票明细表"
    headers1 = [
        "发票号",
        "日期",
        "类型",
        "销方名称",
        "金额",
        "税额",
        "价税合计",
        "费用科目",
        "报销人",
        "置信度",
        "状态",
        "异常原因",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for inv in invoices:
        status = "✓" if inv.anomaly_flag == "normal" else "⚠️"
        ws1.append(
            [
                inv.invoice_no or "",
                str(inv.invoice_date) if inv.invoice_date else "",
                inv.invoice_type or "",
                inv.seller_name or "",
                inv.amount,
                inv.tax_amount,
                inv.total_amount,
                inv.expense_category or "",
                inv.reimbursement_person or "",
                f"{inv.confidence:.0%}" if inv.confidence else "",
                status,
                inv.anomaly_reason or "",
            ]
        )

    # Sheet 2: 汇总表
    ws2 = wb.create_sheet("汇总表")
    headers2 = ["费用科目", "发票数量", "合计金额", "合计税额"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for item in summary:
        ws2.append([item.category, item.count, item.amount, item.tax_amount])

    # Sheet 3: 异常清单
    ws3 = wb.create_sheet("异常清单")
    headers3 = ["发票号", "销方名称", "金额", "异常原因", "原图路径"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for inv in anomalies:
        ws3.append(
            [
                inv.invoice_no or "",
                inv.seller_name or "",
                inv.total_amount,
                inv.anomaly_reason or "",
                inv.image_path or "",
            ]
        )

    # Sheet 4: 凭证导入模板
    ws4 = wb.create_sheet("凭证导入模板")
    headers4 = [
        "编制日期",
        "凭证类型",
        "凭证序号",
        "凭证号",
        "制单人",
        "附件张数",
        "会计年度",
        "科目编码",
        "科目名称",
        "凭证摘要",
        "借贷方向",
        "金额",
        "币种",
        "汇率",
        "原币金额",
        "数量",
        "单价",
        "结算方式名称",
        "结算日期",
        "结算票号",
        "业务日期",
        "员工编号",
        "员工姓名",
        "往来单位编号",
        "往来单位名称",
        "货品编号",
        "货品名称",
        "部门名称",
        "项目名称",
    ]
    ws4.append(headers4)
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for v in vouchers:
        # 确保 None 值输出为空字符串
        ws4.append(
            [
                v.编制日期 or "",
                v.凭证类型 or "",
                v.凭证序号 if v.凭证序号 is not None else "",
                v.凭证号 or "",
                v.制单人 or "",
                v.附件张数 if v.附件张数 is not None else 0,
                v.会计年度 or "",
                v.科目编码 or "",
                v.科目名称 or "",
                v.凭证摘要 or "",
                v.借贷方向 or "",
                v.金额 if v.金额 is not None else 0,
                v.币种 or "人民币",
                v.汇率 if v.汇率 is not None else 1,
                v.原币金额 if v.原币金额 is not None else 0,
                v.数量 if v.数量 is not None else "",
                v.单价 if v.单价 is not None else "",
                v.结算方式名称 or "",
                v.结算日期 or "",
                v.结算票号 or "",
                v.业务日期 or "",
                v.员工编号 or "",
                v.员工姓名 or "",
                v.往来单位编号 or "",
                v.往来单位名称 or "",
                v.货品编号 or "",
                v.货品名称 or "",
                v.部门名称 or "",
                v.项目名称 or "",
            ]
        )

    # 调整列宽
    for ws in [ws1, ws2, ws3, ws4]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
